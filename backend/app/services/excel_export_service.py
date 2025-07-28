from typing import Dict, List, Optional, Any
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, DataBar, DataBarRule
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint

from app.utils.building_type_display import get_display_building_type
from app.utils.formatting import (
    format_currency, format_currency_compact, format_percentage,
    format_square_feet, format_cost_per_sf, format_quantity,
    format_unit_cost, format_line_item
)
from app.utils.grouping import group_similar_items, format_grouped_export
from app.services.executive_summary_service import executive_summary_service


class ExcelExportService:
    """Service for generating professional Excel exports with formulas"""
    
    def __init__(self):
        self._init_styles()
    
    def _init_styles(self):
        """Initialize reusable styles"""
        self.styles = {
            'header': {
                'font': Font(bold=True, size=14, color="FFFFFF"),
                'fill': PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid"),
                'alignment': Alignment(horizontal="center", vertical="center"),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'subheader': {
                'font': Font(bold=True, size=12),
                'fill': PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
                'alignment': Alignment(horizontal="left", vertical="center")
            },
            'title': {
                'font': Font(bold=True, size=18, color="1976D2"),
                'alignment': Alignment(horizontal="center", vertical="center")
            },
            'currency': {
                'number_format': '$#,##0.00',
                'alignment': Alignment(horizontal="right")
            },
            'percentage': {
                'number_format': '0.0%',
                'alignment': Alignment(horizontal="center")
            },
            'total': {
                'font': Font(bold=True, size=12),
                'fill': PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
                'border': Border(top=Side(style='double'))
            }
        }
    
    def generate_excel_report(self, project_data: Dict, include_markups: bool = True) -> io.BytesIO:
        """Generate a comprehensive Excel report with multiple sheets"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Generate executive summary data
        executive_summary = executive_summary_service.generate_executive_summary(project_data)
        
        # Create sheets
        self._create_executive_summary_sheet(wb, project_data, executive_summary)
        self._create_summary_sheet(wb, project_data)
        self._create_detailed_breakdown_sheet(wb, project_data, include_markups)
        self._create_grouped_breakdown_sheet(wb, project_data)  # New grouped sheet
        self._create_trade_packages_sheet(wb, project_data)
        self._create_analysis_sheet(wb, project_data)
        self._create_project_info_sheet(wb, project_data)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_summary_sheet(self, wb: Workbook, project_data: Dict):
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary")
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "Construction Cost Estimate - Executive Summary"
        ws['A1'].font = self.styles['title']['font']
        ws['A1'].alignment = self.styles['title']['alignment']
        
        # Project Info
        row = 3
        ws[f'A{row}'] = "Project Information"
        ws[f'A{row}'].font = self.styles['subheader']['font']
        
        request_data = project_data.get('request_data', {})
        info_items = [
            ('Project Name:', project_data.get('project_name', 'N/A')),
            ('Building Type:', get_display_building_type(request_data)),
            ('Location:', request_data.get('location', 'N/A')),
            ('Square Footage:', f"{request_data.get('square_footage', 0):,} SF"),
            ('Number of Floors:', request_data.get('num_floors', 1)),
            ('Date Generated:', datetime.now().strftime('%B %d, %Y'))
        ]
        
        row += 1
        for label, value in info_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Cost Summary
        row += 2
        ws[f'A{row}'] = "Cost Summary"
        ws[f'A{row}'].font = self.styles['subheader']['font']
        
        # Add headers
        row += 1
        headers = ['Category', 'Base Cost', 'Markup', 'Subtotal', '% of Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
        
        # Add category data with formulas
        start_data_row = row + 1
        categories = project_data.get('categories', [])
        
        for idx, category in enumerate(categories):
            row += 1
            ws[f'A{row}'] = category['name']
            
            # Check if we have markup details
            if 'markup_details' in category:
                base_cost = category['base_subtotal']
                markup = category['markup_details']['total_markup']
                subtotal = category['subtotal_with_markup']
            else:
                base_cost = category.get('subtotal', 0)
                markup = 0
                subtotal = base_cost
            
            # Use clean formatting
            ws[f'B{row}'] = base_cost
            ws[f'C{row}'] = markup
            ws[f'D{row}'] = f'=B{row}+C{row}'  # Formula for subtotal
            ws[f'E{row}'] = f'=D{row}/$D${row + len(categories) + 1}'  # Percentage formula
            
            # Apply currency format
            for col in ['B', 'C', 'D']:
                ws[f'{col}{row}'].number_format = self.styles['currency']['number_format']
            ws[f'E{row}'].number_format = self.styles['percentage']['number_format']
        
        # Add totals row
        total_row = row + 1
        ws[f'A{total_row}'] = "SUBTOTAL"
        ws[f'B{total_row}'] = f'=SUM(B{start_data_row}:B{row})'
        ws[f'C{total_row}'] = f'=SUM(C{start_data_row}:C{row})'
        ws[f'D{total_row}'] = f'=SUM(D{start_data_row}:D{row})'
        ws[f'E{total_row}'] = "100.0%"
        
        # Apply total styling
        for col in range(1, 6):
            cell = ws.cell(row=total_row, column=col)
            cell.font = self.styles['total']['font']
            cell.fill = self.styles['total']['fill']
        
        # Add contingency and final total
        contingency_percent = project_data.get('contingency_percentage', 10)
        row = total_row + 2
        
        ws[f'A{row}'] = f"Contingency ({contingency_percent}%)"
        ws[f'D{row}'] = f'=D{total_row}*{contingency_percent/100}'
        ws[f'D{row}'].number_format = self.styles['currency']['number_format']
        
        row += 1
        ws[f'A{row}'] = "TOTAL PROJECT COST"
        ws[f'D{row}'] = f'=D{total_row}+D{row-1}'
        ws[f'D{row}'].number_format = self.styles['currency']['number_format']
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'D{row}'].font = Font(bold=True, size=14)
        
        # Add cost per SF
        row += 2
        ws[f'A{row}'] = "Cost per Square Foot:"
        ws[f'B{row}'] = f'=D{row-2}/{request_data.get("square_footage", 1)}'
        ws[f'B{row}'].number_format = self.styles['currency']['number_format']
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        # Add a pie chart
        if len(categories) > 0:
            pie = PieChart()
            labels = Reference(ws, min_col=1, min_row=start_data_row, max_row=row-6)
            data = Reference(ws, min_col=4, min_row=start_data_row-1, max_row=row-6)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Cost Distribution by Category"
            pie.height = 10
            pie.width = 15
            ws.add_chart(pie, "G3")
        
        # Adjust column widths
        self._adjust_column_widths(ws, {
            'A': 25,
            'B': 15,
            'C': 15,
            'D': 15,
            'E': 12
        })
    
    def _create_executive_summary_sheet(self, wb: Workbook, project_data: Dict, executive_summary: Dict):
        """Create professional executive summary sheet"""
        ws = wb.create_sheet("Executive Overview", 0)  # Insert as first sheet
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "Executive Summary"
        ws['A1'].font = Font(bold=True, size=20, color="1976D2")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Project Overview Section
        row = 3
        ws[f'A{row}'] = "PROJECT OVERVIEW"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="1976D2")
        row += 2
        
        overview = executive_summary['project_overview']
        for key, value in overview.items():
            ws[f'A{row}'] = key.replace('_', ' ').title() + ':'
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Cost Summary Section
        row += 2
        ws[f'A{row}'] = "COST SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="1976D2")
        row += 2
        
        cost_summary = executive_summary['cost_summary']
        for key, value in cost_summary.items():
            ws[f'A{row}'] = key.replace('_', ' ').title() + ':'
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            if 'cost' in key.lower():
                ws[f'B{row}'].font = Font(size=12)
            row += 1
        
        # Major Systems Breakdown
        row += 2
        ws[f'A{row}'] = "MAJOR SYSTEMS BREAKDOWN"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="1976D2")
        row += 2
        
        # Headers for systems
        headers = ['System', 'Cost', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        row += 1
        for system in executive_summary['major_systems']:
            ws[f'A{row}'] = system['system']
            ws[f'B{row}'] = system['cost']
            ws[f'C{row}'] = system['percentage']
            
            # Alternate row coloring
            if row % 2 == 0:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            row += 1
        
        # Confidence Assessment
        row += 2
        ws[f'A{row}'] = "CONFIDENCE ASSESSMENT"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="1976D2")
        row += 2
        
        confidence = executive_summary['confidence_assessment']
        ws[f'A{row}'] = "Overall Confidence:"
        ws[f'B{row}'] = f"{confidence['overall_confidence']} ({confidence['confidence_level']})"
        ws[f'A{row}'].font = Font(bold=True)
        
        # Style confidence cell based on level
        if confidence['confidence_level'] == "High":
            ws[f'B{row}'].font = Font(color="2E7D32", bold=True)
        elif confidence['confidence_level'] == "Medium":
            ws[f'B{row}'].font = Font(color="F57C00", bold=True)
        else:
            ws[f'B{row}'].font = Font(color="D32F2F", bold=True)
        
        # Key Assumptions
        row += 3
        ws[f'A{row}'] = "KEY ASSUMPTIONS"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="1976D2")
        row += 1
        
        for assumption in executive_summary['key_assumptions'][:5]:
            row += 1
            ws[f'A{row}'] = f"• {assumption}"
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        
        # Adjust column widths
        self._adjust_column_widths(ws, {
            'A': 35,
            'B': 25,
            'C': 15,
            'D': 15,
            'E': 15,
            'F': 15
        })
        
        # Set row heights for better readability
        for r in range(1, row + 1):
            ws.row_dimensions[r].height = 20
    
    def _create_grouped_breakdown_sheet(self, wb: Workbook, project_data: Dict):
        """Create grouped breakdown sheet with similar items together"""
        ws = wb.create_sheet("Grouped Systems")
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = "Systems Grouped by Type"
        ws['A1'].font = self.styles['title']['font']
        ws['A1'].alignment = self.styles['title']['alignment']
        
        # Extract all items for grouping
        all_items = []
        for category in project_data.get('categories', []):
            for system in category.get('systems', []):
                item = {
                    'description': system.get('name', ''),
                    'quantity': system.get('quantity', 0),
                    'unit': system.get('unit', ''),
                    'unit_cost': system.get('unit_cost', 0),
                    'total_cost': system.get('total_cost', 0),
                    'trade': category.get('name', 'General'),
                    'confidence': system.get('confidence_score', 95)
                }
                all_items.append(item)
        
        # Group items
        grouped_data = group_similar_items(all_items, group_by='type')
        
        # Headers
        row = 3
        headers = ['Group/System', 'Quantity', 'Unit', 'Unit Cost', 'Total Cost', 'Trade', 'Confidence']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
        
        row += 1
        
        # Add grouped data
        for group_summary in grouped_data['summary']:
            # Group header
            ws[f'A{row}'] = group_summary['group']
            ws[f'E{row}'] = format_currency(group_summary['total_cost'])
            ws[f'G{row}'] = f"{group_summary['item_count']} items"
            
            # Style group header
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="42A5F5", end_color="42A5F5", fill_type="solid")
            
            row += 1
            
            # Add items in group
            for item in group_summary['items']:
                ws[f'A{row}'] = f"  {item['description']}"  # Indent
                ws[f'B{row}'] = item['quantity']
                ws[f'C{row}'] = item['unit']
                ws[f'D{row}'] = item['unit_cost']
                ws[f'E{row}'] = item['total_cost']
                ws[f'F{row}'] = item['trade']
                ws[f'G{row}'] = format_percentage(item['confidence'], 0)
                
                # Apply formats
                ws[f'B{row}'].number_format = '#,##0.00'
                ws[f'D{row}'].number_format = self.styles['currency']['number_format']
                ws[f'E{row}'].number_format = self.styles['currency']['number_format']
                
                # Alternate row colors
                if row % 2 == 0:
                    for col in range(1, 8):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
                        )
                
                row += 1
            
            # Group subtotal
            ws[f'A{row}'] = f"{group_summary['group']} Total"
            ws[f'E{row}'] = format_currency(group_summary['total_cost'])
            
            # Style subtotal
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, italic=True)
                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            
            row += 2  # Add spacing
        
        # Grand total
        total_cost = sum(item['total_cost'] for item in all_items)
        ws[f'A{row}'] = "GRAND TOTAL"
        ws[f'E{row}'] = format_currency(total_cost)
        
        # Style grand total
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True, size=14)
            cell.fill = self.styles['total']['fill']
            cell.border = Border(top=Side(style='double'))
        
        # Adjust column widths
        self._adjust_column_widths(ws, {
            'A': 40,
            'B': 12,
            'C': 10,
            'D': 15,
            'E': 15,
            'F': 20,
            'G': 12
        })
    
    def _create_detailed_breakdown_sheet(self, wb: Workbook, project_data: Dict, include_markups: bool):
        """Create detailed cost breakdown sheet"""
        ws = wb.create_sheet("Detailed Breakdown")
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = "Detailed Cost Breakdown"
        ws['A1'].font = self.styles['title']['font']
        ws['A1'].alignment = self.styles['title']['alignment']
        
        # Headers
        row = 3
        headers = ['Category', 'System', 'Quantity', 'Unit', 'Unit Cost', 'Base Cost', 'Markup', 'Total Cost']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
        
        # Add data
        row += 1
        start_row = row
        
        for category in project_data.get('categories', []):
            category_start = row
            
            for system in category.get('systems', []):
                ws[f'A{row}'] = category['name']
                ws[f'B{row}'] = system['name']
                ws[f'C{row}'] = system['quantity']
                ws[f'D{row}'] = system['unit']
                ws[f'E{row}'] = system['unit_cost']
                ws[f'F{row}'] = f'=C{row}*E{row}'  # Base cost formula
                
                # Add markup if available
                if include_markups and 'markup_details' in category:
                    markup_percent = (category['markup_details']['overhead_percent'] + 
                                    category['markup_details']['profit_percent']) / 100
                    ws[f'G{row}'] = f'=F{row}*{markup_percent}'
                else:
                    ws[f'G{row}'] = 0
                
                ws[f'H{row}'] = f'=F{row}+G{row}'  # Total cost formula
                
                # Apply number formats
                ws[f'C{row}'].number_format = '#,##0.00'
                ws[f'E{row}'].number_format = self.styles['currency']['number_format']
                ws[f'F{row}'].number_format = self.styles['currency']['number_format']
                ws[f'G{row}'].number_format = self.styles['currency']['number_format']
                ws[f'H{row}'].number_format = self.styles['currency']['number_format']
                
                row += 1
            
            # Add category subtotal
            if row > category_start:
                ws[f'A{row}'] = f"{category['name']} Subtotal"
                ws[f'F{row}'] = f'=SUM(F{category_start}:F{row-1})'
                ws[f'G{row}'] = f'=SUM(G{category_start}:G{row-1})'
                ws[f'H{row}'] = f'=SUM(H{category_start}:H{row-1})'
                
                # Apply subtotal styling
                for col in range(1, 9):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                row += 1
        
        # Add grand total
        total_row = row + 1
        ws[f'A{total_row}'] = "GRAND TOTAL"
        ws[f'F{total_row}'] = f'=SUMIF(A{start_row}:A{row},"*Subtotal",F{start_row}:F{row})'
        ws[f'G{total_row}'] = f'=SUMIF(A{start_row}:A{row},"*Subtotal",G{start_row}:G{row})'
        ws[f'H{total_row}'] = f'=SUMIF(A{start_row}:A{row},"*Subtotal",H{start_row}:H{row})'
        
        # Apply grand total styling
        for col in range(1, 9):
            cell = ws.cell(row=total_row, column=col)
            cell.font = self.styles['total']['font']
            cell.fill = self.styles['total']['fill']
            cell.border = self.styles['total']['border']
        
        # Create a table
        table_ref = f"A3:H{row-1}"
        table = Table(displayName="DetailedBreakdown", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                             showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Adjust column widths
        self._adjust_column_widths(ws, {
            'A': 20,
            'B': 35,
            'C': 12,
            'D': 10,
            'E': 12,
            'F': 15,
            'G': 12,
            'H': 15
        })
    
    def _create_trade_packages_sheet(self, wb: Workbook, project_data: Dict):
        """Create trade packages summary sheet"""
        ws = wb.create_sheet("Trade Packages")
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = "Trade Package Summary"
        ws['A1'].font = self.styles['title']['font']
        ws['A1'].alignment = self.styles['title']['alignment']
        
        # Get trade summaries
        trade_summaries = project_data.get('trade_summaries', {})
        
        # Headers
        row = 3
        headers = ['Trade', 'Base Cost', 'Markup %', 'Total Cost', '% of Project']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
        
        # Add trade data
        row += 1
        start_row = row
        total_project_cost = project_data.get('total_cost', 0)
        
        for trade_name, trade_data in trade_summaries.items():
            if isinstance(trade_data, dict) and 'total' in trade_data:
                ws[f'A{row}'] = trade_data.get('displayName', trade_name.title())
                
                # Calculate base cost and markup
                trade_total = trade_data['total']
                
                # Find markup percentage for this trade
                markup_percent = 0
                for category in project_data.get('categories', []):
                    if category['name'].lower() in trade_name.lower():
                        if 'markup_details' in category:
                            markup_percent = category['markup_details']['effective_markup_percent']
                        break
                
                base_cost = trade_total / (1 + markup_percent/100) if markup_percent > 0 else trade_total
                
                ws[f'B{row}'] = base_cost
                ws[f'C{row}'] = markup_percent / 100
                ws[f'D{row}'] = trade_total
                ws[f'E{row}'] = f'=D{row}/{total_project_cost}'
                
                # Apply formats
                ws[f'B{row}'].number_format = self.styles['currency']['number_format']
                ws[f'C{row}'].number_format = self.styles['percentage']['number_format']
                ws[f'D{row}'].number_format = self.styles['currency']['number_format']
                ws[f'E{row}'].number_format = self.styles['percentage']['number_format']
                
                row += 1
        
        # Add total
        total_row = row + 1
        ws[f'A{total_row}'] = "TOTAL"
        ws[f'B{total_row}'] = f'=SUM(B{start_row}:B{row-1})'
        ws[f'D{total_row}'] = f'=SUM(D{start_row}:D{row-1})'
        ws[f'E{total_row}'] = "100.0%"
        
        # Apply total styling
        for col in range(1, 6):
            cell = ws.cell(row=total_row, column=col)
            cell.font = self.styles['total']['font']
            cell.fill = self.styles['total']['fill']
        
        # Add notes
        notes_row = total_row + 3
        ws[f'A{notes_row}'] = "Notes:"
        ws[f'A{notes_row}'].font = Font(bold=True)
        
        notes = [
            "• Base costs exclude overhead and profit markups",
            "• Markup percentages include both overhead and profit",
            "• Trade packages can be exported individually from the project detail view",
            "• Percentages are calculated based on total project cost including contingency"
        ]
        
        for note in notes:
            notes_row += 1
            ws[f'A{notes_row}'] = note
            ws[f'A{notes_row}'].font = Font(size=10, italic=True)
        
        # Adjust column widths
        self._adjust_column_widths(ws, {
            'A': 25,
            'B': 15,
            'C': 12,
            'D': 15,
            'E': 15
        })
    
    def _create_analysis_sheet(self, wb: Workbook, project_data: Dict):
        """Create analysis sheet with metrics and comparisons"""
        ws = wb.create_sheet("Analysis")
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "Project Analysis & Metrics"
        ws['A1'].font = self.styles['title']['font']
        ws['A1'].alignment = self.styles['title']['alignment']
        
        # Key Metrics
        row = 3
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = self.styles['subheader']['font']
        
        request_data = project_data.get('request_data', {})
        sqft = request_data.get('square_footage', 1)
        total_cost = project_data.get('total_cost', 0)
        
        metrics = [
            ('Total Project Cost:', total_cost, self.styles['currency']['number_format']),
            ('Cost per Square Foot:', total_cost / sqft if sqft > 0 else 0, self.styles['currency']['number_format']),
            ('Total Square Footage:', sqft, '#,##0'),
            ('Number of Floors:', request_data.get('num_floors', 1), '#,##0'),
            ('Average Cost per Floor:', total_cost / request_data.get('num_floors', 1), self.styles['currency']['number_format']),
        ]
        
        row += 1
        for label, value, format_str in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].number_format = format_str
            row += 1
        
        # Cost breakdown by percentage
        row += 2
        ws[f'A{row}'] = "Cost Distribution Analysis"
        ws[f'A{row}'].font = self.styles['subheader']['font']
        
        row += 1
        ws[f'A{row}'] = "Category"
        ws[f'B{row}'] = "Cost"
        ws[f'C{row}'] = "% of Total"
        ws[f'D{row}'] = "Visual"
        
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        
        # Add category analysis
        categories = project_data.get('categories', [])
        row += 1
        start_analysis_row = row
        
        for category in sorted(categories, key=lambda x: x.get('subtotal', 0), reverse=True):
            ws[f'A{row}'] = category['name']
            
            if 'subtotal_with_markup' in category:
                cost = category['subtotal_with_markup']
            else:
                cost = category.get('subtotal', 0)
            
            ws[f'B{row}'] = cost
            ws[f'C{row}'] = cost / total_cost if total_cost > 0 else 0
            
            ws[f'B{row}'].number_format = self.styles['currency']['number_format']
            ws[f'C{row}'].number_format = self.styles['percentage']['number_format']
            
            row += 1
        
        # Add data bars to percentage column
        if row > start_analysis_row:
            rule = DataBarRule(
                start_type='num', start_value=0,
                end_type='num', end_value=1,
                color="FF638EC6"
            )
            ws.conditional_formatting.add(f'C{start_analysis_row}:C{row-1}', rule)
        
        # Regional comparison (if available)
        if 'regional_multiplier' in project_data:
            row += 2
            ws[f'A{row}'] = "Regional Cost Factors"
            ws[f'A{row}'].font = self.styles['subheader']['font']
            
            row += 1
            ws[f'A{row}'] = "Location:"
            ws[f'B{row}'] = request_data.get('location', 'N/A')
            
            row += 1
            ws[f'A{row}'] = "Regional Multiplier:"
            ws[f'B{row}'] = project_data.get('regional_multiplier', 1.0)
            ws[f'B{row}'].number_format = '0.00'
        
        # Adjust column widths
        self._adjust_column_widths(ws, {
            'A': 30,
            'B': 15,
            'C': 15,
            'D': 20
        })
    
    def _create_project_info_sheet(self, wb: Workbook, project_data: Dict):
        """Create project information sheet"""
        ws = wb.create_sheet("Project Info")
        
        # Title
        ws.merge_cells('A1:B1')
        ws['A1'] = "Project Information & Specifications"
        ws['A1'].font = self.styles['title']['font']
        ws['A1'].alignment = self.styles['title']['alignment']
        
        # Project details
        row = 3
        request_data = project_data.get('request_data', {})
        
        sections = [
            ("Basic Information", [
                ('Project ID:', project_data.get('project_id', 'N/A')),
                ('Project Name:', project_data.get('project_name', 'N/A')),
                ('Created Date:', project_data.get('created_at', datetime.now().isoformat()).split('T')[0]),
                ('Building Type:', get_display_building_type(request_data)),
                ('Project Type:', request_data.get('project_type', 'N/A').replace('_', ' ').title()),
            ]),
            ("Location & Size", [
                ('Location:', request_data.get('location', 'N/A')),
                ('Climate Zone:', request_data.get('climate_zone', 'N/A').replace('_', ' ').title() if request_data.get('climate_zone') else 'N/A'),
                ('Square Footage:', format_square_feet(request_data.get('square_footage', 0))),
                ('Number of Floors:', request_data.get('num_floors', 1)),
                ('Ceiling Height:', f"{request_data.get('ceiling_height', 9)} ft"),
            ]),
            ("Special Requirements", [
                ('Description:', request_data.get('special_requirements', 'None specified')),
                ('Occupancy Type:', request_data.get('occupancy_type', 'N/A').title()),
            ])
        ]
        
        for section_title, items in sections:
            ws[f'A{row}'] = section_title
            ws[f'A{row}'].font = self.styles['subheader']['font']
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            for label, value in items:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            row += 1
        
        # Add markup settings if available
        if 'markup_summary' in project_data:
            ws[f'A{row}'] = "Markup Settings"
            ws[f'A{row}'].font = self.styles['subheader']['font']
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            markup_summary = project_data['markup_summary']
            ws[f'A{row}'] = "Average Markup:"
            ws[f'B{row}'] = f"{markup_summary.get('average_markup_percent', 0):.1f}%"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            ws[f'A{row}'] = "Show in PDF:"
            ws[f'B{row}'] = "Yes" if markup_summary.get('show_in_pdf', True) else "No"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Adjust column widths
        self._adjust_column_widths(ws, {
            'A': 25,
            'B': 40
        })
    
    def _adjust_column_widths(self, ws, widths: Dict[str, int]):
        """Adjust column widths"""
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width
    
    def generate_trade_excel(self, trade_data: Dict, trade_name: str, project_data: Dict) -> io.BytesIO:
        """Generate Excel export for a specific trade package"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"{trade_name.title()} Package"
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"{trade_name.title()} Trade Package"
        ws['A1'].font = self.styles['title']['font']
        ws['A1'].alignment = self.styles['title']['alignment']
        
        # Project info
        row = 3
        ws['A3'] = "Project:"
        ws['B3'] = project_data.get('project_name', 'N/A')
        ws['A4'] = "Location:"
        ws['B4'] = project_data.get('request_data', {}).get('location', 'N/A')
        ws['A5'] = "Date:"
        ws['B5'] = datetime.now().strftime('%B %d, %Y')
        
        for r in range(3, 6):
            ws[f'A{r}'].font = Font(bold=True)
        
        # Systems breakdown
        row = 7
        headers = ['System', 'Quantity', 'Unit', 'Unit Cost', 'Total Cost', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
        
        # Add systems
        row += 1
        start_row = row
        
        for system in trade_data.get('systems', []):
            ws[f'A{row}'] = system['name']
            ws[f'B{row}'] = system['quantity']
            ws[f'C{row}'] = system['unit']
            ws[f'D{row}'] = system['unit_cost']
            ws[f'E{row}'] = f'=B{row}*D{row}'
            ws[f'F{row}'] = system.get('specifications', {}).get('note', '')
            
            # Apply formats
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'D{row}'].number_format = self.styles['currency']['number_format']
            ws[f'E{row}'].number_format = self.styles['currency']['number_format']
            
            row += 1
        
        # Add total
        total_row = row + 1
        ws[f'A{total_row}'] = "TOTAL"
        ws[f'E{total_row}'] = f'=SUM(E{start_row}:E{row-1})'
        
        # Apply total styling
        for col in range(1, 7):
            cell = ws.cell(row=total_row, column=col)
            cell.font = self.styles['total']['font']
            cell.fill = self.styles['total']['fill']
        
        # Add markup section if applicable
        if 'markup_details' in trade_data:
            markup_row = total_row + 2
            ws[f'A{markup_row}'] = "Markup Details"
            ws[f'A{markup_row}'].font = self.styles['subheader']['font']
            
            markup_row += 1
            ws[f'A{markup_row}'] = "Base Cost:"
            ws[f'B{markup_row}'] = f'=E{total_row}'
            
            markup_row += 1
            ws[f'A{markup_row}'] = f"Overhead ({trade_data['markup_details']['overhead_percent']}%):"
            ws[f'B{markup_row}'] = f'=B{markup_row-1}*{trade_data["markup_details"]["overhead_percent"]/100}'
            
            markup_row += 1
            ws[f'A{markup_row}'] = f"Profit ({trade_data['markup_details']['profit_percent']}%):"
            ws[f'B{markup_row}'] = f'=B{markup_row-2}*{trade_data["markup_details"]["profit_percent"]/100}'
            
            markup_row += 1
            ws[f'A{markup_row}'] = "TOTAL WITH MARKUP:"
            ws[f'B{markup_row}'] = f'=B{markup_row-3}+B{markup_row-2}+B{markup_row-1}'
            
            # Apply formats
            for r in range(total_row + 3, markup_row + 1):
                ws[f'B{r}'].number_format = self.styles['currency']['number_format']
                if r == markup_row:
                    ws[f'A{r}'].font = Font(bold=True, size=12)
                    ws[f'B{r}'].font = Font(bold=True, size=12)
        
        # Adjust column widths
        self._adjust_column_widths(ws, {
            'A': 35,
            'B': 12,
            'C': 10,
            'D': 12,
            'E': 15,
            'F': 30
        })
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output


excel_export_service = ExcelExportService()
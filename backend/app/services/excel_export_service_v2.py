from typing import Dict, List, Optional, Any
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.header_footer import HeaderFooter

from app.utils.building_type_display import get_display_building_type
from app.utils.formatting import (
    format_currency, format_currency_compact, format_percentage,
    format_square_feet, format_cost_per_sf
)
from app.services.executive_summary_service import executive_summary_service


class ProfessionalExcelExportService:
    """Service for generating premium Excel exports with professional formatting"""
    
    # Brand colors
    BRAND_PRIMARY = "667EEA"  # Purple
    BRAND_SECONDARY = "764BA2"  # Dark purple
    BRAND_ACCENT = "10B981"  # Green
    BRAND_WARNING = "F59E0B"  # Orange
    BRAND_DANGER = "EF4444"  # Red
    
    def __init__(self):
        self._init_professional_styles()
    
    def _init_professional_styles(self):
        """Initialize premium styles for $799/month software look"""
        self.styles = {
            'main_title': {
                'font': Font(name="Calibri", bold=True, size=24, color="FFFFFF"),
                'fill': PatternFill(start_color=self.BRAND_PRIMARY, end_color=self.BRAND_SECONDARY, fill_type="solid"),
                'alignment': Alignment(horizontal="center", vertical="center"),
                'border': Border(
                    left=Side(style='medium', color="FFFFFF"),
                    right=Side(style='medium', color="FFFFFF"),
                    top=Side(style='medium', color="FFFFFF"),
                    bottom=Side(style='medium', color="FFFFFF")
                )
            },
            'section_header': {
                'font': Font(name="Calibri", bold=True, size=16, color=self.BRAND_PRIMARY),
                'alignment': Alignment(horizontal="left", vertical="center"),
                'border': Border(bottom=Side(style='thick', color=self.BRAND_PRIMARY))
            },
            'subsection_header': {
                'font': Font(name="Calibri", bold=True, size=14, color="333333"),
                'fill': PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
                'alignment': Alignment(horizontal="left", vertical="center"),
                'border': Border(
                    left=Side(style='thin', color="E5E7EB"),
                    right=Side(style='thin', color="E5E7EB"),
                    top=Side(style='thin', color="E5E7EB"),
                    bottom=Side(style='thin', color="E5E7EB")
                )
            },
            'table_header': {
                'font': Font(name="Calibri", bold=True, size=12, color="FFFFFF"),
                'fill': PatternFill(start_color=self.BRAND_PRIMARY, end_color=self.BRAND_PRIMARY, fill_type="solid"),
                'alignment': Alignment(horizontal="center", vertical="center"),
                'border': Border(
                    left=Side(style='thin', color="FFFFFF"),
                    right=Side(style='thin', color="FFFFFF"),
                    top=Side(style='thin', color="FFFFFF"),
                    bottom=Side(style='thin', color="FFFFFF")
                )
            },
            'currency_cell': {
                'number_format': '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)',
                'alignment': Alignment(horizontal="right", vertical="center"),
                'font': Font(name="Calibri", size=11)
            },
            'percentage_cell': {
                'number_format': '0.0%',
                'alignment': Alignment(horizontal="center", vertical="center"),
                'font': Font(name="Calibri", size=11)
            },
            'total_row': {
                'font': Font(name="Calibri", bold=True, size=12),
                'fill': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
                'border': Border(
                    top=Side(style='double', color="333333"),
                    bottom=Side(style='medium', color="333333")
                )
            },
            'grand_total': {
                'font': Font(name="Calibri", bold=True, size=14, color="FFFFFF"),
                'fill': PatternFill(start_color=self.BRAND_ACCENT, end_color=self.BRAND_ACCENT, fill_type="solid"),
                'alignment': Alignment(horizontal="center", vertical="center"),
                'border': Border(
                    left=Side(style='medium', color="FFFFFF"),
                    right=Side(style='medium', color="FFFFFF"),
                    top=Side(style='medium', color="FFFFFF"),
                    bottom=Side(style='medium', color="FFFFFF")
                )
            },
            'highlight_positive': {
                'fill': PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
                'font': Font(name="Calibri", color="065F46", bold=True)
            },
            'highlight_negative': {
                'fill': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
                'font': Font(name="Calibri", color="991B1B", bold=True)
            }
        }
    
    def generate_professional_excel(self, project_data: Dict, client_name: str = None) -> io.BytesIO:
        """Generate a premium Excel report with professional formatting"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Generate executive summary data
        executive_summary = executive_summary_service.generate_executive_summary(project_data)
        
        # Create sheets in professional order
        self._create_cover_sheet(wb, project_data, client_name)
        self._create_executive_dashboard(wb, project_data, executive_summary)
        self._create_cost_breakdown_sheet(wb, project_data)
        self._create_trade_analysis_sheet(wb, project_data)
        self._create_detailed_systems_sheet(wb, project_data)
        self._create_assumptions_sheet(wb, project_data)
        
        # Apply professional workbook properties
        wb.properties.creator = "SpecSharp Professional"
        wb.properties.company = "SpecSharp - Premium Construction Estimating"
        wb.properties.created = datetime.now()
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_cover_sheet(self, wb: Workbook, project_data: Dict, client_name: str):
        """Create professional cover sheet"""
        ws = wb.create_sheet("Cover", 0)
        
        # Set page margins for professional look
        ws.page_margins = PageMargins(
            left=0.75, right=0.75, top=1.0, bottom=1.0,
            header=0.5, footer=0.5
        )
        
        # Hide gridlines for cleaner look
        ws.sheet_view.showGridLines = False
        
        # Company branding area
        ws.merge_cells('B2:F4')
        ws['B2'] = "SPECSHARP"
        ws['B2'].font = Font(name="Calibri", size=36, bold=True, color=self.BRAND_PRIMARY)
        ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells('B5:F5')
        ws['B5'] = "Professional Construction Cost Estimate"
        ws['B5'].font = Font(name="Calibri", size=16, color="666666")
        ws['B5'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Project title
        ws.merge_cells('B8:F10')
        ws['B8'] = project_data.get('project_name', 'Construction Project')
        ws['B8'].font = Font(name="Calibri", size=24, bold=True)
        ws['B8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['B8'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        ws['B8'].border = Border(
            left=Side(style='medium', color=self.BRAND_PRIMARY),
            right=Side(style='medium', color=self.BRAND_PRIMARY),
            top=Side(style='medium', color=self.BRAND_PRIMARY),
            bottom=Side(style='medium', color=self.BRAND_PRIMARY)
        )
        
        # Key information boxes
        info_boxes = [
            ("TOTAL PROJECT COST", format_currency(project_data.get('total_cost', 0)), self.BRAND_ACCENT),
            ("COST PER SQ FT", f"${project_data.get('cost_per_sqft', 0):.2f}", self.BRAND_PRIMARY),
            ("TOTAL AREA", f"{project_data.get('request_data', {}).get('square_footage', 0):,} SF", self.BRAND_SECONDARY)
        ]
        
        start_row = 13
        for idx, (label, value, color) in enumerate(info_boxes):
            col = 2 + (idx * 2)
            ws.merge_cells(f'{get_column_letter(col)}{start_row}:{get_column_letter(col+1)}{start_row+2}')
            cell = ws.cell(row=start_row, column=col)
            cell.value = f"{label}\n{value}"
            cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='medium', color="FFFFFF"),
                right=Side(style='medium', color="FFFFFF"),
                top=Side(style='medium', color="FFFFFF"),
                bottom=Side(style='medium', color="FFFFFF")
            )
        
        # Client information
        if client_name:
            ws['B18'] = "Prepared for:"
            ws['B18'].font = Font(name="Calibri", size=12, color="666666")
            ws['B19'] = client_name
            ws['B19'].font = Font(name="Calibri", size=16, bold=True)
        
        # Date and metadata
        ws['B22'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws['B22'].font = Font(name="Calibri", size=10, color="999999")
        
        ws['B23'] = f"Valid for: 30 days"
        ws['B23'].font = Font(name="Calibri", size=10, color="999999")
        
        # Set column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Set row heights for better spacing
        ws.row_dimensions[1].height = 30
        for row in range(2, 25):
            ws.row_dimensions[row].height = 25
    
    def _create_executive_dashboard(self, wb: Workbook, project_data: Dict, executive_summary: Dict):
        """Create executive dashboard with charts and key metrics"""
        ws = wb.create_sheet("Executive Dashboard")
        
        # Hide gridlines
        ws.sheet_view.showGridLines = False
        
        # Title
        ws.merge_cells('A1:H2')
        ws['A1'] = "Executive Dashboard"
        ws['A1'].font = self.styles['main_title']['font']
        ws['A1'].fill = self.styles['main_title']['fill']
        ws['A1'].alignment = self.styles['main_title']['alignment']
        ws['A1'].border = self.styles['main_title']['border']
        
        # Key Metrics Section
        ws['A4'] = "KEY METRICS"
        ws['A4'].font = self.styles['section_header']['font']
        for col in range(1, 9):
            ws.cell(row=4, column=col).border = self.styles['section_header']['border']
        
        # Metric cards
        metrics = [
            ("Total Project Cost", format_currency(project_data.get('total_cost', 0)), "Large investment figure"),
            ("Cost per Square Foot", f"${project_data.get('cost_per_sqft', 0):.2f}", "Market competitive rate"),
            ("Total Square Footage", f"{project_data.get('request_data', {}).get('square_footage', 0):,}", "Building size"),
            ("Number of Floors", str(project_data.get('request_data', {}).get('num_floors', 1)), "Vertical complexity"),
            ("Location Factor", self._get_location_factor(project_data), "Regional adjustment"),
            ("Confidence Level", "High (95%)", "Estimate accuracy")
        ]
        
        row = 6
        col = 1
        for metric_name, metric_value, metric_desc in metrics:
            # Metric box
            ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(col+1)}{row+1}')
            cell = ws.cell(row=row, column=col)
            cell.value = metric_value
            cell.font = Font(name="Calibri", size=20, bold=True, color=self.BRAND_PRIMARY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Metric label
            label_cell = ws.cell(row=row+2, column=col)
            ws.merge_cells(f'{get_column_letter(col)}{row+2}:{get_column_letter(col+1)}{row+2}')
            label_cell.value = metric_name
            label_cell.font = Font(name="Calibri", size=11, bold=True)
            label_cell.alignment = Alignment(horizontal="center")
            
            # Metric description
            desc_cell = ws.cell(row=row+3, column=col)
            ws.merge_cells(f'{get_column_letter(col)}{row+3}:{get_column_letter(col+1)}{row+3}')
            desc_cell.value = metric_desc
            desc_cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
            desc_cell.alignment = Alignment(horizontal="center")
            
            # Add border around metric box
            for r in range(row, row+4):
                for c in range(col, col+2):
                    ws.cell(row=r, column=c).border = Border(
                        left=Side(style='thin', color="E5E7EB"),
                        right=Side(style='thin', color="E5E7EB"),
                        top=Side(style='thin', color="E5E7EB"),
                        bottom=Side(style='thin', color="E5E7EB")
                    )
            
            col += 3
            if col > 7:
                col = 1
                row += 5
        
        # Cost Breakdown Chart
        chart_row = row + 2
        ws[f'A{chart_row}'] = "COST DISTRIBUTION"
        ws[f'A{chart_row}'].font = self.styles['section_header']['font']
        for col in range(1, 9):
            ws.cell(row=chart_row, column=col).border = self.styles['section_header']['border']
        
        # Add pie chart
        chart_row += 2
        self._add_cost_pie_chart(ws, project_data, chart_row)
        
        # Trade breakdown table
        table_row = chart_row + 15
        ws[f'A{table_row}'] = "TRADE BREAKDOWN"
        ws[f'A{table_row}'].font = self.styles['section_header']['font']
        for col in range(1, 9):
            ws.cell(row=table_row, column=col).border = self.styles['section_header']['border']
        
        # Add trade table
        table_row += 2
        headers = ['Trade', 'Base Cost', 'Markup', 'Total Cost', '% of Total', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=table_row, column=col, value=header)
            cell.font = self.styles['table_header']['font']
            cell.fill = self.styles['table_header']['fill']
            cell.alignment = self.styles['table_header']['alignment']
            cell.border = self.styles['table_header']['border']
        
        # Add trade data
        categories = project_data.get('categories', [])
        total_cost = project_data.get('total_cost', 0)
        
        for category in categories:
            table_row += 1
            ws[f'A{table_row}'] = category['name']
            ws[f'B{table_row}'] = category.get('base_subtotal', category['subtotal'])
            ws[f'C{table_row}'] = category.get('markup_details', {}).get('total_markup', 0)
            ws[f'D{table_row}'] = category['subtotal']
            ws[f'E{table_row}'] = category['subtotal'] / total_cost if total_cost > 0 else 0
            ws[f'F{table_row}'] = "✓ Included"
            
            # Apply formatting
            ws[f'B{table_row}'].style = 'Currency'
            ws[f'C{table_row}'].style = 'Currency'
            ws[f'D{table_row}'].style = 'Currency'
            ws[f'E{table_row}'].number_format = '0.0%'
            ws[f'F{table_row}'].font = Font(color=self.BRAND_ACCENT)
            
            # Alternate row shading
            if table_row % 2 == 0:
                for col in range(1, 7):
                    ws.cell(row=table_row, column=col).fill = PatternFill(
                        start_color="F9FAFB", end_color="F9FAFB", fill_type="solid"
                    )
        
        # Add total row
        table_row += 1
        ws[f'A{table_row}'] = "TOTAL"
        ws[f'D{table_row}'] = f'=SUM(D{table_row-len(categories)}:D{table_row-1})'
        ws[f'E{table_row}'] = 1.0
        
        for col in range(1, 7):
            cell = ws.cell(row=table_row, column=col)
            cell.font = self.styles['total_row']['font']
            cell.fill = self.styles['total_row']['fill']
            cell.border = self.styles['total_row']['border']
        
        # Adjust column widths
        column_widths = {'A': 20, 'B': 18, 'C': 15, 'D': 18, 'E': 12, 'F': 15, 'G': 15, 'H': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Add SpecSharp branding at bottom
        branding_row = table_row + 3
        ws.merge_cells(f'A{branding_row}:H{branding_row}')
        ws[f'A{branding_row}'] = "Powered by SpecSharp - Professional Estimates in 90 Seconds - specsharp.ai"
        ws[f'A{branding_row}'].font = Font(name="Calibri", size=11, bold=True, color=self.BRAND_PRIMARY)
        ws[f'A{branding_row}'].alignment = Alignment(horizontal="center", vertical="center")
        ws[f'A{branding_row}'].fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
        
        # Add clickable link
        ws[f'A{branding_row}'].hyperlink = "https://specsharp.ai?ref=excel"
        
        # Freeze panes for scrolling
        ws.freeze_panes = 'A5'
    
    def _create_cost_breakdown_sheet(self, wb: Workbook, project_data: Dict):
        """Create detailed cost breakdown sheet"""
        ws = wb.create_sheet("Cost Breakdown")
        
        # Hide gridlines
        ws.sheet_view.showGridLines = False
        
        # Title
        ws.merge_cells('A1:H2')
        ws['A1'] = "Detailed Cost Breakdown"
        ws['A1'].font = self.styles['main_title']['font']
        ws['A1'].fill = self.styles['main_title']['fill']
        ws['A1'].alignment = self.styles['main_title']['alignment']
        
        # Summary section
        row = 4
        ws[f'A{row}'] = "COST SUMMARY"
        ws[f'A{row}'].font = self.styles['section_header']['font']
        
        row += 2
        summary_data = [
            ('Subtotal (All Trades):', project_data.get('subtotal', 0)),
            ('Contingency:', project_data.get('contingency_amount', 0)),
            ('Total Project Cost:', project_data.get('total_cost', 0))
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'C{row}'] = value
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'C{row}'].style = 'Currency'
            ws[f'C{row}'].font = Font(size=12)
            
            if 'Total Project' in label:
                ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
                ws[f'C{row}'].font = Font(bold=True, size=14, color="FFFFFF")
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = self.styles['grand_total']['fill']
            
            row += 1
        
        # Detailed breakdown by category
        row += 2
        ws[f'A{row}'] = "BREAKDOWN BY TRADE"
        ws[f'A{row}'].font = self.styles['section_header']['font']
        
        row += 2
        
        # Create detailed table
        headers = ['Trade', 'System', 'Quantity', 'Unit', 'Unit Cost', 'Total Cost', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles['table_header']['font']
            cell.fill = self.styles['table_header']['fill']
            cell.alignment = self.styles['table_header']['alignment']
            cell.border = self.styles['table_header']['border']
        
        row += 1
        
        # Process each category
        for category in project_data.get('categories', []):
            category_start_row = row
            
            # Merge cells for trade name
            ws.merge_cells(f'A{row}:A{row + len(category.get("systems", [])) - 1}')
            ws[f'A{row}'] = category['name']
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws[f'A{row}'].alignment = Alignment(vertical='center')
            ws[f'A{row}'].fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
            
            # Add systems
            for system in category.get('systems', []):
                ws[f'B{row}'] = system['name']
                ws[f'C{row}'] = system['quantity']
                ws[f'D{row}'] = system['unit']
                ws[f'E{row}'] = system['unit_cost']
                ws[f'F{row}'] = system['total_cost']
                ws[f'G{row}'] = f"Confidence: {system.get('confidence_score', 95)}%"
                
                # Apply formatting
                ws[f'E{row}'].style = 'Currency'
                ws[f'F{row}'].style = 'Currency'
                
                # Alternate row shading
                if row % 2 == 0:
                    for col in range(2, 8):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="F9FAFB", end_color="F9FAFB", fill_type="solid"
                        )
                
                row += 1
            
            # Trade subtotal
            ws[f'E{row}'] = "Subtotal:"
            ws[f'F{row}'] = category['subtotal']
            ws[f'E{row}'].font = Font(bold=True)
            ws[f'F{row}'].font = Font(bold=True)
            ws[f'F{row}'].style = 'Currency'
            ws[f'F{row}'].border = Border(top=Side(style='thin'))
            
            row += 2
        
        # Adjust column widths
        column_widths = {
            'A': 18, 'B': 35, 'C': 12, 'D': 10, 
            'E': 15, 'F': 18, 'G': 25, 'H': 15
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze panes
        ws.freeze_panes = 'A11'
        
    def _add_cost_pie_chart(self, ws, project_data: Dict, start_row: int):
        """Add professional pie chart for cost distribution"""
        # Prepare data for chart
        categories = project_data.get('categories', [])
        
        # Add data for chart (hidden)
        data_col = 10  # Column J
        ws.cell(row=start_row, column=data_col, value="Category")
        ws.cell(row=start_row, column=data_col+1, value="Amount")
        
        for idx, category in enumerate(categories):
            ws.cell(row=start_row+idx+1, column=data_col, value=category['name'])
            ws.cell(row=start_row+idx+1, column=data_col+1, value=category['subtotal'])
        
        # Create pie chart
        pie = PieChart()
        pie.title = "Cost Distribution by Trade"
        pie.style = 10  # Professional style
        
        # Add data to chart
        labels = Reference(ws, min_col=data_col, min_row=start_row+1, max_row=start_row+len(categories))
        data = Reference(ws, min_col=data_col+1, min_row=start_row, max_row=start_row+len(categories))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        
        # Professional colors
        colors = [self.BRAND_PRIMARY, self.BRAND_SECONDARY, self.BRAND_ACCENT, 
                  "FFA500", "FF6B6B", "4ECDC4", "95E1D3"]
        
        for idx, point in enumerate(pie.series[0].data_points):
            point.graphicalProperties.solidFill = colors[idx % len(colors)]
        
        # Position and size
        pie.height = 10
        pie.width = 15
        pie.legend.position = 'r'
        
        ws.add_chart(pie, f'A{start_row}')
        
        # Hide data columns
        ws.column_dimensions[get_column_letter(data_col)].hidden = True
        ws.column_dimensions[get_column_letter(data_col+1)].hidden = True
    
    def _get_location_factor(self, project_data: Dict) -> str:
        """Get location factor for display"""
        location = project_data.get('request_data', {}).get('location', '')
        
        # High cost locations
        high_cost = ['San Francisco', 'New York', 'Boston', 'Seattle']
        low_cost = ['Phoenix', 'Dallas', 'Houston', 'Atlanta']
        
        for city in high_cost:
            if city.lower() in location.lower():
                return "High (+15-45%)"
        
        for city in low_cost:
            if city.lower() in location.lower():
                return "Low (-10-20%)"
        
        return "Standard (±5%)"
    
    def _add_branding_footer(self, ws, row: int, start_col: str = 'A', end_col: str = 'H'):
        """Add SpecSharp branding footer to any worksheet"""
        ws.merge_cells(f'{start_col}{row}:{end_col}{row}')
        branding_cell = ws[f'{start_col}{row}']
        branding_cell.value = "Powered by SpecSharp • Professional Construction Estimates in 90 Seconds • Visit specsharp.ai"
        branding_cell.font = Font(name="Calibri", size=10, bold=True, italic=True, color=self.BRAND_PRIMARY)
        branding_cell.alignment = Alignment(horizontal="center", vertical="center")
        branding_cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        branding_cell.hyperlink = "https://specsharp.ai?ref=excel-export"
        
        # Add border
        thin_border = Side(style='thin', color="E5E7EB")
        for col_num in range(ord(start_col) - ord('A'), ord(end_col) - ord('A') + 1):
            cell = ws.cell(row=row, column=col_num + 1)
            cell.border = Border(
                top=thin_border,
                bottom=thin_border,
                left=thin_border if col_num == ord(start_col) - ord('A') else None,
                right=thin_border if col_num == ord(end_col) - ord('A') else None
            )
    
    def _create_detailed_systems_sheet(self, wb: Workbook, project_data: Dict):
        """Create detailed systems breakdown with professional formatting"""
        ws = wb.create_sheet("Detailed Systems")
        
        # Hide gridlines
        ws.sheet_view.showGridLines = False
        
        # Title
        ws.merge_cells('A1:G2')
        ws['A1'] = "Detailed Systems Breakdown"
        ws['A1'].font = self.styles['main_title']['font']
        ws['A1'].fill = self.styles['main_title']['fill']
        ws['A1'].alignment = self.styles['main_title']['alignment']
        
        row = 4
        
        # Process each category
        for category in project_data.get('categories', []):
            # Category header
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = category['name'].upper()
            ws[f'A{row}'].font = self.styles['section_header']['font']
            ws[f'A{row}'].fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
            
            row += 2
            
            # Table headers
            headers = ['System Description', 'Quantity', 'Unit', 'Unit Cost', 'Total Cost', 'Notes', 'Confidence']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.styles['table_header']['font']
                cell.fill = self.styles['table_header']['fill']
                cell.alignment = self.styles['table_header']['alignment']
            
            row += 1
            data_start_row = row
            
            # Add systems
            for system in category.get('systems', []):
                ws[f'A{row}'] = system['name']
                ws[f'B{row}'] = system['quantity']
                ws[f'C{row}'] = system['unit']
                ws[f'D{row}'] = system['unit_cost']
                ws[f'E{row}'] = system['total_cost']
                ws[f'F{row}'] = "Included in estimate"
                ws[f'G{row}'] = f"{system.get('confidence_score', 95)}%"
                
                # Apply formatting
                ws[f'D{row}'].style = 'Currency'
                ws[f'E{row}'].style = 'Currency'
                
                # Alternate row colors
                if row % 2 == 0:
                    for col in range(1, 8):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="F9FAFB", end_color="F9FAFB", fill_type="solid"
                        )
                
                row += 1
            
            # Category total
            ws[f'A{row}'] = f"{category['name']} Total"
            ws[f'E{row}'] = f'=SUM(E{data_start_row}:E{row-1})'
            
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.font = self.styles['total_row']['font']
                cell.fill = self.styles['total_row']['fill']
                cell.border = Border(top=Side(style='double'))
            
            row += 3
        
        # Adjust column widths
        column_widths = {'A': 40, 'B': 12, 'C': 10, 'D': 15, 'E': 18, 'F': 25, 'G': 12}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def _create_trade_analysis_sheet(self, wb: Workbook, project_data: Dict):
        """Create trade analysis sheet with charts"""
        ws = wb.create_sheet("Trade Analysis")
        
        # Hide gridlines
        ws.sheet_view.showGridLines = False
        
        # Title
        ws.merge_cells('A1:F2')
        ws['A1'] = "Trade Analysis & Comparison"
        ws['A1'].font = self.styles['main_title']['font']
        ws['A1'].fill = self.styles['main_title']['fill']
        ws['A1'].alignment = self.styles['main_title']['alignment']
        
        # Add bar chart placeholder
        ws['A4'] = "Trade Cost Comparison"
        ws['A4'].font = self.styles['section_header']['font']
        
        # Note about chart
        ws['A6'] = "See Executive Dashboard for visual representation"
        ws['A6'].font = Font(italic=True, color="666666")
        
        # Trade comparison table
        ws['A8'] = "TRADE COMPARISON TABLE"
        ws['A8'].font = self.styles['section_header']['font']
        
        # Headers
        headers = ['Trade', 'Base Cost', 'Markup %', 'Markup $', 'Total Cost', 'Industry Avg']
        row = 10
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles['table_header']['font']
            cell.fill = self.styles['table_header']['fill']
            cell.alignment = self.styles['table_header']['alignment']
        
        # Add trade data with industry comparisons
        industry_averages = {
            'Structural': 0.20,
            'Mechanical': 0.25,
            'Electrical': 0.15,
            'Plumbing': 0.10,
            'Finishes': 0.30
        }
        
        total_cost = project_data.get('total_cost', 0)
        
        for category in project_data.get('categories', []):
            row += 1
            trade_name = category['name']
            base_cost = category.get('base_subtotal', category['subtotal'])
            markup = category.get('markup_details', {}).get('total_markup', 0)
            markup_percent = (markup / base_cost * 100) if base_cost > 0 else 0
            total = category['subtotal']
            industry_avg = industry_averages.get(trade_name, 0.20) * total_cost
            
            ws[f'A{row}'] = trade_name
            ws[f'B{row}'] = base_cost
            ws[f'C{row}'] = markup_percent / 100
            ws[f'D{row}'] = markup
            ws[f'E{row}'] = total
            ws[f'F{row}'] = industry_avg
            
            # Apply formatting
            ws[f'B{row}'].style = 'Currency'
            ws[f'C{row}'].number_format = '0.0%'
            ws[f'D{row}'].style = 'Currency'
            ws[f'E{row}'].style = 'Currency'
            ws[f'F{row}'].style = 'Currency'
            
            # Highlight if significantly different from industry average
            variance = (total - industry_avg) / industry_avg if industry_avg > 0 else 0
            if abs(variance) > 0.20:  # 20% variance
                color = self.styles['highlight_negative']['fill'] if variance > 0 else self.styles['highlight_positive']['fill']
                ws[f'E{row}'].fill = color
    
    def _create_assumptions_sheet(self, wb: Workbook, project_data: Dict):
        """Create assumptions and notes sheet"""
        ws = wb.create_sheet("Assumptions & Notes")
        
        # Hide gridlines
        ws.sheet_view.showGridLines = False
        
        # Title
        ws.merge_cells('A1:F2')
        ws['A1'] = "Project Assumptions & Important Notes"
        ws['A1'].font = self.styles['main_title']['font']
        ws['A1'].fill = self.styles['main_title']['fill']
        ws['A1'].alignment = self.styles['main_title']['alignment']
        
        row = 4
        
        # Key Assumptions
        ws[f'A{row}'] = "KEY ASSUMPTIONS"
        ws[f'A{row}'].font = self.styles['section_header']['font']
        
        row += 2
        assumptions = [
            "All pricing based on current market conditions as of estimate date",
            "Labor rates calculated using regional prevailing wages",
            "Material costs include standard grade materials unless specified",
            "Estimate assumes normal working conditions and hours",
            "No allowance for overtime or accelerated schedule",
            "Site is assumed to be clear and ready for construction",
            "Normal soil conditions assumed - no rock excavation included",
            "All work to be performed during regular business hours"
        ]
        
        for assumption in assumptions:
            ws[f'A{row}'] = f"• {assumption}"
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        # Exclusions
        row += 2
        ws[f'A{row}'] = "EXCLUSIONS"
        ws[f'A{row}'].font = self.styles['section_header']['font']
        
        row += 2
        exclusions = [
            "Land acquisition costs",
            "Financing charges and interest",
            "Professional fees (architectural, engineering, legal)",
            "Permits and government fees",
            "FF&E (Furniture, Fixtures & Equipment)",
            "Site work beyond building footprint",
            "Utility connection fees",
            "Testing and commissioning",
            "Contingency for unknown conditions"
        ]
        
        for exclusion in exclusions:
            ws[f'A{row}'] = f"• {exclusion}"
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
        
        # Terms
        row += 2
        ws[f'A{row}'] = "TERMS & CONDITIONS"
        ws[f'A{row}'].font = self.styles['section_header']['font']
        
        row += 2
        ws.merge_cells(f'A{row}:F{row+3}')
        ws[f'A{row}'] = (
            "This estimate is valid for 30 days from the date shown. Pricing is subject to change based on "
            "market conditions, material availability, and labor rates. This is a budgetary estimate only "
            "and should not be considered a firm bid. Final pricing subject to detailed plans and specifications."
        )
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical="top")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 80


# Export instance
excel_export_service_v2 = ProfessionalExcelExportService()
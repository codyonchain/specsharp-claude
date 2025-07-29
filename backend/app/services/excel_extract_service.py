"""
Excel Extract Service for Subcontractor Scopes
Generates Excel files with trade-specific scope items without pricing
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List


class ExcelExtractService:
    """Service for extracting trade-specific scopes for subcontractors"""
    
    def __init__(self):
        # Brand colors
        self.primary_color = "667EEA"
        self.secondary_color = "F7FAFC"
        self.header_color = "2D3748"
        
    def generate_subcontractor_scope(
        self, 
        trade_data: Dict[str, Any], 
        trade_name: str,
        project: Any
    ) -> io.BytesIO:
        """Generate Excel file with trade scope for subcontractors"""
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_cover_sheet(wb, trade_name, project)
        self._create_scope_sheet(wb, trade_data, trade_name)
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    def _create_cover_sheet(self, wb: Workbook, trade_name: str, project: Any):
        """Create professional cover page"""
        ws = wb.create_sheet("Cover Page")
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        
        # Title
        ws['B2'] = f"{trade_name.upper()} SCOPE OF WORK"
        ws['B2'].font = Font(size=24, bold=True, color=self.primary_color)
        ws['B2'].alignment = Alignment(horizontal='left')
        
        # Project name
        ws['B4'] = project.name
        ws['B4'].font = Font(size=18, bold=True)
        
        # Date
        ws['B6'] = f"Prepared: {datetime.now().strftime('%B %d, %Y')}"
        ws['B6'].font = Font(size=12, color="666666")
        
        # Project Information Section
        ws['B9'] = "PROJECT INFORMATION"
        ws['B9'].font = Font(size=14, bold=True, color=self.header_color)
        
        # Project details
        details = [
            ("Project Name:", project.name),
            ("Location:", project.location),
            ("Building Type:", project.building_type or "Commercial"),
            ("Square Footage:", f"{project.square_footage:,} SF"),
            ("Number of Floors:", str(project.num_floors)),
        ]
        
        row = 11
        for label, value in details:
            ws[f'B{row}'] = label
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'] = value
            ws[f'C{row}'].font = Font(size=11)
            row += 1
        
        # Instructions section
        ws[f'B{row+2}'] = "INSTRUCTIONS FOR BIDDERS"
        ws[f'B{row+2}'].font = Font(size=14, bold=True, color=self.header_color)
        
        instructions = [
            "1. Review all scope items carefully",
            "2. Provide unit pricing for each line item",
            "3. Include all labor, materials, and equipment in your pricing",
            "4. Indicate any exclusions or clarifications",
            "5. Submit completed bid by the due date"
        ]
        
        row += 4
        for instruction in instructions:
            ws[f'B{row}'] = instruction
            ws[f'B{row}'].font = Font(size=11)
            row += 1
        
        # Footer with stronger branding
        ws[f'B{row+3}'] = "Powered by SpecSharp"
        ws[f'B{row+3}'].font = Font(size=12, bold=True, color=self.primary_color)
        ws[f'B{row+4}'] = "Professional Construction Estimating in 90 Seconds"
        ws[f'B{row+4}'].font = Font(size=10, italic=True, color="666666")
        ws[f'B{row+5}'] = "Visit specsharp.ai to create your own estimates"
        ws[f'B{row+5}'].font = Font(size=10, bold=True, color=self.primary_color)
    
    def _create_scope_sheet(self, wb: Workbook, trade_data: Dict[str, Any], trade_name: str):
        """Create detailed scope sheet for the trade"""
        ws = wb.create_sheet(f"{trade_name} Scope")
        
        # Set column widths
        column_widths = {
            'A': 8,   # Item #
            'B': 50,  # Description
            'C': 15,  # Quantity
            'D': 12,  # Unit
            'E': 20,  # Unit Price (Contractor)
            'F': 20,  # Total Price (Contractor)
            'G': 30,  # Notes/Clarifications
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Header
        ws['B1'] = f"{trade_name.upper()} - DETAILED SCOPE"
        ws['B1'].font = Font(size=16, bold=True, color=self.primary_color)
        ws.merge_cells('B1:F1')
        
        # Column headers
        headers = [
            "Item #",
            "Scope Description",
            "Quantity",
            "Unit",
            "Unit Price",
            "Total Price",
            "Notes/Clarifications"
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color=self.header_color, end_color=self.header_color, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Write scope items
        row_num = 4
        item_num = 1
        
        systems = trade_data.get('systems', [])
        
        # Sort systems by name for consistency
        sorted_systems = sorted(systems, key=lambda x: x.get('name', ''))
        
        for system in sorted_systems:
            # Skip if no quantity
            if not system.get('quantity', 0):
                continue
            
            ws.cell(row=row_num, column=1, value=item_num).border = border
            ws.cell(row=row_num, column=2, value=system.get('name', '')).border = border
            ws.cell(row=row_num, column=3, value=f"{system.get('quantity', 0):,.2f}").border = border
            ws.cell(row=row_num, column=4, value=system.get('unit', '')).border = border
            
            # Empty cells for contractor to fill
            ws.cell(row=row_num, column=5, value="").border = border  # Unit Price
            ws.cell(row=row_num, column=5).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            
            ws.cell(row=row_num, column=6, value="").border = border  # Total Price
            ws.cell(row=row_num, column=6).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            
            ws.cell(row=row_num, column=7, value="").border = border  # Notes
            
            # Add formula for total calculation
            ws.cell(row=row_num, column=6).value = f"=C{row_num}*E{row_num}"
            
            row_num += 1
            item_num += 1
        
        # Add subtotal row
        row_num += 1
        ws.cell(row=row_num, column=2, value="SUBTOTAL").font = Font(bold=True)
        ws.cell(row=row_num, column=6, value=f"=SUM(F4:F{row_num-2})").font = Font(bold=True)
        ws.cell(row=row_num, column=6).border = Border(top=Side(style='double'))
        
        # Add notes section
        row_num += 3
        ws.cell(row=row_num, column=1, value="NOTES:")
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.merge_cells(f'B{row_num}:G{row_num+5}')
        ws.cell(row=row_num, column=2).border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # SpecSharp Branding Footer
        row_num += 8
        ws.merge_cells(f'A{row_num}:G{row_num}')
        ws.cell(row=row_num, column=1, value="Scope extracted using SpecSharp • Get your own account at specsharp.ai")
        ws.cell(row=row_num, column=1).font = Font(size=11, bold=True, color=self.primary_color)
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=1).fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
        
        row_num += 1
        ws.merge_cells(f'A{row_num}:G{row_num}')
        ws.cell(row=row_num, column=1, value="Create professional estimates in 90 seconds • Join 1,000+ contractors")
        ws.cell(row=row_num, column=1).font = Font(size=10, italic=True, color="666666")
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
        
        # Freeze panes
        ws.freeze_panes = 'A4'


# Create service instance
excel_extract_service = ExcelExtractService()